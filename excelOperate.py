from openpyxl import load_workbook
import os

charNumLimit = 6000
separator = '∞'


class operateExecl:
    def __init__(self, file):
        self.file = file
        self.__openFlag = False
        self.openFlag = self.__openFlag
        self.__handle = None
        self.__ws = None
        self.title = None
        self.max_row = None
        self.max_column = None
        self.sheetnames = None

    def openExcel(self):
        if not self.__openFlag:
            self.__handle = load_workbook(self.file)
            self.sheetnames = self.__handle.sheetnames
            self.__ws = self.__handle[self.sheetnames[0]]
            self.title = self.__ws.title
            self.max_row = self.__ws.max_row
            self.max_column = self.__ws.max_column
            self.openFlag = self.__openFlag = True
        else:
            self.openFlag = True

    def closeExcel(self):
        self.__handle.close()
        self.openFlag = self.__openFlag = False

    def insertRow(self, idx, amount=1):
        for i in range(amount):
            self.__ws.insert_row(idx)
            self.max_row += 1

    def insertColumn(self, idx, amount=1):
        for i in range(amount):
            self.__ws.insert_cols(idx)
            self.max_column += 1

    def writeExcel(self, row, column, value, sheetName=""):
        if sheetName != "":
            self.__ws = self.__handle[sheetName]
        self.__ws.cell(row, column).value = value

    def readExcel(self, row, column, sheetName=""):
        if sheetName != "":
            self.__ws = self.__handle[sheetName]
        return self.__ws.cell(row, column).value

    def saveExcel(self):
        self.__handle.save(self.file)
